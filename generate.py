#!/usr/bin/env python3
"""
TGT AOA — Management Dashboard Generator
=========================================
Reads TGT Master File_26-27.xlsx and writes index.html.

The design (CSS, layout, sections) is locked in this script.
Only the numbers change when the Master File is updated.

Usage:
    python3 generate.py

Then commit and push:
    git add index.html
    git commit -m "Dashboard update"
    git push

Or just run:  ./update.sh
"""

import openpyxl
import sys
from datetime import datetime, date
from collections import Counter
from pathlib import Path

# ── CONFIG ───────────────────────────────────────────────────────────────────
MASTER_FILE = Path(
    "/Users/abhaykapoor/Library/Mobile Documents/"
    "com~apple~CloudDocs/Claude_TGT/TGT Master File_26-27.xlsx"
)
OUTPUT_FILE = Path(__file__).parent / "index.html"
TODAY = date.today()
WEEK_NUM = TODAY.isocalendar()[1]
FY_LABEL = "FY 2026-27"
# ─────────────────────────────────────────────────────────────────────────────


# ── HELPERS ───────────────────────────────────────────────────────────────────

def ind(n):
    """Indian number format: 1723089 → 17,23,089. Returns — for None/zero."""
    if n is None:
        return "—"
    try:
        neg = float(n) < 0
        n = abs(int(round(float(n))))
    except (TypeError, ValueError):
        return "—"
    if n == 0:
        return "—"
    s = str(n)
    if len(s) <= 3:
        r = s
    else:
        r = s[-3:]
        s = s[:-3]
        while s:
            chunk = s[-2:] if len(s) >= 2 else s
            r = chunk + "," + r
            s = s[:-2]
    return f"({r})" if neg else r


def parse_date(val):
    """Parse MyGate date strings ('DD-MM-YY, HH:MM AM/PM') or datetime objects."""
    if not val:
        return TODAY
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d-%m-%y, %I:%M %p", "%d-%m-%Y, %I:%M %p", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return TODAY


def age_cls(days):
    if days >= 100:
        return "age-critical"
    if days >= 30:
        return "age-high"
    if days >= 9:
        return "age-med"
    return "age-low"


def status_pill(status):
    m = {
        "In Progress": "pill-prog",
        "Reopened":    "pill-reopen",
        "On Hold":     "pill-hold",
    }
    cls = m.get(status, "pill-prog")
    return f'<span class="badge-pill {cls}">{status}</span>'


def pri_pill(p):
    m = {"High": "pill-high", "Medium": "pill-med", "Low": "pill-low"}
    cls = m.get(p, "pill-med")
    return f'<span class="badge-pill {cls}">{p}</span>' if p else "—"


# ── DATA EXTRACTION ───────────────────────────────────────────────────────────

def get_tickets(wb):
    ws = wb["Helpdesk Tickets_Raw"]
    rows = list(ws.iter_rows(values_only=True))
    h = {v: i for i, v in enumerate(rows[0]) if v}
    open_st = {"In Progress", "Reopened", "On Hold"}
    out = []
    for r in rows[1:]:
        st = str(r[h["Status"]] or "").strip()
        if st not in open_st:
            continue
        cd = parse_date(r[h["Created Date"]])
        lu = parse_date(r[h["Last Update"]])
        out.append({
            "id":       r[h["Id"]],
            "category": str(r[h["Category"]] or "Other"),
            "flat":     str(r[h["Flat"]] or "").replace("TOWER ", ""),
            "subject":  str(r[h["Subject"]] or "")[:65],
            "status":   st,
            "created":  cd,
            "last_upd": lu,
            "age":      (TODAY - cd).days,
            "idle":     (TODAY - lu).days,
        })
    return out


def get_financials(wb):
    ws = wb["Income Expense Sheet"]
    rows = list(ws.iter_rows(values_only=True))
    # Row index 3 = header: "Row Labels" | "May" | "June" | "Grand Total"
    header = rows[3]
    months = [
        (i, v) for i, v in enumerate(header)
        if v and v not in ("Row Labels", "Grand Total")
    ]
    if not months:
        return {}, "Unknown"
    col, label = months[-1]   # most recent month column

    cr, dr = {}, {}
    section = None
    for r in rows[4:]:
        lbl, val = r[0], r[col]
        if lbl == "CR":
            section = "cr"
        elif lbl == "DR":
            section = "dr"
        elif lbl == "Grand Total":
            break
        elif lbl and section == "cr":
            cr[str(lbl)] = float(val or 0)
        elif lbl and section == "dr":
            dr[str(lbl)] = float(val or 0)

    total_cr = sum(cr.values())
    total_dr = sum(dr.values())
    return {
        "cr": cr, "dr": dr,
        "total_cr": total_cr,
        "total_dr": total_dr,
        "deficit":  total_cr - total_dr,
    }, str(label)


def get_payments(wb):
    ws = wb["Payment Tracking Sheet"]
    rows = list(ws.iter_rows(values_only=True))
    vendors, totals = [], {}
    for r in rows:
        if len(r) < 18:
            continue
        name = r[9]
        if name and isinstance(name, str) and name != "Vendor Name":
            b  = float(r[11] or 0)
            py = float(r[15] or 0)
            pd = float(r[16] or 0)
            pn = float(r[17] or 0)
            if b > 0 or py > 0:
                vendors.append({
                    "vendor":  name,
                    "purpose": str(r[10] or ""),
                    "billed":  b,
                    "payable": py,
                    "paid":    pd,
                    "pending": pn,
                })
        # Totals row: col 9 empty, col 15 is a big number
        elif (r[9] is None and r[15]
              and isinstance(r[15], (int, float))
              and float(r[15]) > 10000
              and not totals):
            totals = {
                "payable": float(r[15] or 0),
                "paid":    float(r[16] or 0),
                "pending": float(r[17] or 0),
            }
    return vendors, totals


def get_dues(wb):
    ws = wb["Dues Report Raw"]
    rows = list(ws.iter_rows(values_only=True))
    res, com = [], []
    for r in rows[1:]:
        if not r[0]:
            continue
        unit    = str(r[0])
        penalty = float(r[4] or 0)
        ledger  = float(r[5] or 0)
        total   = float(r[6] or 0)
        if unit.startswith("Non Members"):
            com.append({
                "name":   unit.replace("Non Members-", ""),
                "ledger": ledger,
                "total":  total,
            })
        else:
            res.append({
                "unit":    unit.replace("TOWER ", ""),
                "owner":   str(r[1] or "—"),
                "tenant":  str(r[2] or "—"),
                "penalty": penalty,
                "ledger":  ledger,
                "total":   total,
            })
    return res, com


def get_collection(wb):
    ws = wb["Collection Report Raw"]
    rows = list(ws.iter_rows(values_only=True))
    counts, amounts = Counter(), {}
    for r in rows[1:]:
        ds  = str(r[1] or "")
        amt = float(r[10] or 0)
        # Expect "2026-06-..." format
        if len(ds) >= 7 and ds[4] == "-" and ds[7] == "-":
            mk = ds[:7]   # "2026-06"
            counts[mk] += 1
            amounts[mk] = amounts.get(mk, 0) + amt
    sorted_months = sorted(counts.keys(), reverse=True)
    return [(m, counts[m], amounts.get(m, 0)) for m in sorted_months[:2]]


def get_actions(wb):
    ws = wb["Action Tracker"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows:
        if not any(r):
            continue
        is_numbered   = r[0] and isinstance(r[0], int)
        is_unnumbered = (r[0] is None and r[2] and r[3]
                         and isinstance(r[3], str) and r[9])
        if not (is_numbered or is_unnumbered):
            continue
        num      = r[0] if is_numbered else "—"
        date_val = r[1]
        date_str = (date_val.strftime("%d-%b-%y")
                    if isinstance(date_val, datetime) else str(date_val or ""))
        out.append({
            "num":         num,
            "date":        date_str,
            "week":        str(r[2] or ""),
            "category":    str(r[3] or ""),
            "description": str(r[4] or ""),
            "assigned":    str(r[6] or ""),
            "priority":    str(r[7] or ""),
            "status":      str(r[9] or ""),
        })
    return out


def get_commitments(wb):
    ws = wb["Approved Commitments"]
    rows = list(ws.iter_rows(values_only=True))
    items, total = [], 0
    for r in rows:
        if not (r[0] and isinstance(r[0], int)):
            continue
        status = str(r[7] or "")
        if "Pending" not in status:
            continue
        amt      = float(r[5] or 0)
        date_val = r[1]
        date_str = (date_val.strftime("%d-%b-%y")
                    if isinstance(date_val, datetime) else str(date_val or ""))
        items.append({
            "num":      r[0],
            "date":     date_str,
            "vendor":   str(r[2] or ""),
            "desc":     str(r[3] or ""),
            "category": str(r[4] or ""),
            "amount":   amt,
            "month":    str(r[6] or ""),
        })
        total += amt
    return items, total


# ── HTML FRAGMENT BUILDERS ────────────────────────────────────────────────────

BAR_COLORS = {
    "Seepage":                 "#c62828",
    "Plumbing":                "#1565c0",
    "Electrical":              "#f57f17",
    "Housekeeping":            "#2e7d32",
    "Ground Floor Common Area":"#2e7d32",
    "Other":                   "#455a64",
}

CR_ORDER = [
    "Member Fee", "Rental Income", "FD Interest",
    "Promotional Income", "Coaching Revenue", "Miscllaneous", "Vendor Electricity",
]
DR_ORDER = [
    "Facilities Maintenance", "Security Management", "Water Bill",
    "Electricity Bill", "Water Management", "Other Repairs", "Diesel",
    "Landscaping", "Pest Control", "Accountant Fee", "TDS Payment",
    "GST Payment", "Miscllaneous", "Event Related Expense", "Bank Charges",
]


def _ordered_rows(data, order):
    html, shown = "", set()
    for key in order:
        for k, v in data.items():
            if key.lower() in k.lower() and k not in shown:
                html += f"<tr><td>{k}</td><td class='r'>{ind(v)}</td></tr>\n"
                shown.add(k)
    for k, v in data.items():
        if k not in shown:
            html += f"<tr><td>{k}</td><td class='r'>{ind(v)}</td></tr>\n"
    return html


def ticket_table_rows(tickets, mode="age"):
    html = ""
    for t in tickets:
        if mode == "age":
            metric_val = t["age"]
            metric_date = t["created"].strftime("%d-%b-%y")
            last_col_header = "Created"
        else:
            metric_val = t["idle"]
            metric_date = t["last_upd"].strftime("%d-%b-%y")
            last_col_header = "Last Update"
        html += (
            f"<tr>"
            f"<td>{t['id']}</td>"
            f"<td>{t['flat']}</td>"
            f"<td>{t['category']}</td>"
            f"<td>{t['subject']}</td>"
            f"<td>{status_pill(t['status'])}</td>"
            f"<td class='r'><span class='age-pill {age_cls(metric_val)}'>{metric_val}d</span></td>"
            f"<td>{metric_date}</td>"
            f"</tr>\n"
        )
    return html


def on_hold_rows(tickets):
    html = ""
    for t in tickets:
        html += (
            f"<tr>"
            f"<td>{t['id']}</td>"
            f"<td>{t['flat']}</td>"
            f"<td>{t['category']}</td>"
            f"<td>{t['subject']}</td>"
            f"<td class='r'><span class='age-pill {age_cls(t['age'])}'>{t['age']}d</span></td>"
            f"</tr>\n"
        )
    return html


def cat_bars_html(cat_counts, total):
    html = '<div class="cat-bars">\n'
    for cat, cnt in cat_counts.most_common():
        pct = round(cnt / total * 100, 1) if total else 0
        color = BAR_COLORS.get(cat, "#546e7a")
        html += (
            f'<div class="cat-row">'
            f'<div class="cat-name">{cat}</div>'
            f'<div class="bar-wrap"><div class="bar-fill" style="width:{pct}%;background:{color};">'
            f'<span class="bar-label">{pct}%</span></div></div>'
            f'<div class="cat-count">{cnt}</div>'
            f'</div>\n'
        )
    html += "</div>\n"
    return html


def res_due_rows(residential):
    html = ""
    tp, tl, tt = 0, 0, 0
    for r in residential:
        html += (
            f"<tr>"
            f"<td>{r['unit']}</td>"
            f"<td>{r['owner']}</td>"
            f"<td>{r['tenant']}</td>"
            f"<td class='r'>{ind(r['penalty']) if r['penalty'] else '—'}</td>"
            f"<td class='r'>{ind(r['ledger'])}</td>"
            f"<td class='r'>{ind(r['total'])}</td>"
            f"</tr>\n"
        )
        tp += r["penalty"]; tl += r["ledger"]; tt += r["total"]
    html += (
        f"<tr class='total-row'>"
        f"<td colspan='3'>TOTAL</td>"
        f"<td class='r'>{ind(tp)}</td>"
        f"<td class='r'>{ind(tl)}</td>"
        f"<td class='r'>{ind(tt)}</td>"
        f"</tr>\n"
    )
    return html, tt


def com_due_rows(commercial):
    html = ""
    grand = 0
    for c in commercial:
        html += (
            f"<tr>"
            f"<td>{c['name']}</td>"
            f"<td class='r'>{ind(c['ledger'])}</td>"
            f"<td class='r'>{ind(c['total'])}</td>"
            f"</tr>\n"
        )
        grand += c["total"]
    html += (
        f"<tr class='total-row'>"
        f"<td>TOTAL</td>"
        f"<td class='r'>{ind(grand)}</td>"
        f"<td class='r'>{ind(grand)}</td>"
        f"</tr>\n"
    )
    return html, grand


def vendor_rows_html(vendors):
    html = ""
    for v in vendors:
        pn    = float(v.get("pending") or 0)
        cls   = " class='alert-row'" if pn > 0 else ""
        pn_s  = ind(pn) if pn > 0 else "—"
        html += (
            f"<tr{cls}>"
            f"<td>{v['vendor']}</td>"
            f"<td>{v['purpose']}</td>"
            f"<td class='r'>{ind(v['billed'])}</td>"
            f"<td class='r'>{ind(v['payable'])}</td>"
            f"<td class='r'>{ind(v['paid'])}</td>"
            f"<td class='r'>{pn_s}</td>"
            f"</tr>\n"
        )
    return html


def action_rows_html(actions):
    html = ""
    for a in actions:
        is_open = "Open" in a["status"] or "🟡" in a["status"]
        cls     = " class='alert-row'" if is_open else ""
        html += (
            f"<tr{cls}>"
            f"<td>{a['num']}</td>"
            f"<td>{a['week']}</td>"
            f"<td>{a['category']}</td>"
            f"<td>{a['description']}</td>"
            f"<td>{a['assigned']}</td>"
            f"<td>{pri_pill(a['priority'])}</td>"
            f"<td class='c'>{a['status']}</td>"
            f"</tr>\n"
        )
    return html


def commitment_rows_html(items):
    html = ""
    for it in items:
        html += (
            f"<tr>"
            f"<td>{it['num']}</td>"
            f"<td>{it['date']}</td>"
            f"<td>{it['vendor']}</td>"
            f"<td>{it['desc']}</td>"
            f"<td>{it['category']}</td>"
            f"<td class='r'>{ind(it['amount'])}</td>"
            f"<td>{it['month']}</td>"
            f"</tr>\n"
        )
    return html


def collection_kpis(collection):
    html = ""
    for m, cnt, amt in collection:
        try:
            label = datetime.strptime(m, "%Y-%m").strftime("%b %Y")
        except ValueError:
            label = m
        html += (
            f'<div class="kpi">'
            f'<div class="kpi-val">{cnt}</div>'
            f'<div class="kpi-lbl">{label} Receipts</div>'
            f'</div>\n'
            f'<div class="kpi">'
            f'<div class="kpi-val">&#8377;{ind(amt)}</div>'
            f'<div class="kpi-lbl">{label} Total Collections</div>'
            f'</div>\n'
        )
    return html


# ── CSS ───────────────────────────────────────────────────────────────────────

CSS = """\
:root {
  --navy:#1a3a5c; --navy-mid:#234f7a; --navy-light:#2d6a9f;
  --green:#2e7d32; --green-light:#e8f5e9;
  --amber:#e65100; --amber-light:#fff3e0;
  --red:#c62828; --red-light:#ffebee;
  --grey-light:#f5f7f9; --white:#fff;
  --border:#dde3e8; --text:#1a2733; --text-mid:#455a64; --text-light:#90a4ae;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#eef1f5;color:var(--text);font-size:13.5px;line-height:1.5}
.hdr{background:var(--navy);color:#fff;padding:18px 28px;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:100;box-shadow:0 2px 8px rgba(0,0,0,.25)}
.hdr-title{font-size:17px;font-weight:700;letter-spacing:.3px}
.hdr-sub{font-size:11.5px;color:#90caf9;margin-top:2px}
.hdr-meta{text-align:right;font-size:11px;color:#b0bec5;line-height:1.8}
.page{max-width:1080px;margin:0 auto;padding:22px 20px 40px}
.section{background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:18px;overflow:hidden}
.sec-hdr{background:var(--navy);color:#fff;padding:11px 18px;font-size:12.5px;font-weight:700;letter-spacing:.6px;text-transform:uppercase;display:flex;justify-content:space-between;align-items:center}
.sec-hdr .sec-tag{font-size:10px;font-weight:500;background:rgba(255,255,255,.15);padding:2px 8px;border-radius:20px;letter-spacing:.3px;text-transform:none}
.sec-body{padding:16px 18px}
.kpi-row{display:grid;gap:12px;margin-bottom:14px}
.kpi-row-3{grid-template-columns:repeat(3,1fr)}
.kpi-row-4{grid-template-columns:repeat(4,1fr)}
.kpi{background:var(--grey-light);border-radius:8px;padding:12px 14px;border-left:4px solid var(--navy-mid)}
.kpi-val{font-size:22px;font-weight:800;color:var(--navy);line-height:1.1}
.kpi-lbl{font-size:10.5px;color:var(--text-mid);margin-top:4px;font-weight:500}
.kpi.red{border-left-color:var(--red);background:var(--red-light)}
.kpi.red .kpi-val{color:var(--red)}
.kpi.amber{border-left-color:var(--amber);background:var(--amber-light)}
.kpi.amber .kpi-val{color:var(--amber)}
.kpi.green{border-left-color:var(--green);background:var(--green-light)}
.kpi.green .kpi-val{color:var(--green)}
.alert{border-radius:7px;padding:10px 14px;margin-bottom:12px;font-size:12px;line-height:1.6}
.alert.red{background:var(--red-light);border-left:4px solid var(--red);color:var(--red)}
.alert.amber{background:var(--amber-light);border-left:4px solid var(--amber);color:#7f3700}
.alert strong{font-weight:700}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:var(--navy);color:#fff;padding:7px 10px;text-align:left;font-size:10.5px;font-weight:600;letter-spacing:.4px;white-space:nowrap}
th.r,td.r{text-align:right}
th.c,td.c{text-align:center}
td{padding:6px 10px;border-bottom:1px solid var(--border);vertical-align:top}
tr:nth-child(even) td{background:#f8fafc}
tr:hover td{background:#edf4fb}
.tbl-wrap{overflow-x:auto}
.tbl-scroll{max-height:360px;overflow-y:auto}
.tbl-scroll table thead th{position:sticky;top:0;z-index:5}
.badge-pill{display:inline-block;padding:2px 8px;border-radius:12px;font-size:10px;font-weight:700;white-space:nowrap}
.pill-prog{background:#e3f2fd;color:#1565c0}
.pill-hold{background:#f3e5f5;color:#7b1fa2}
.pill-reopen{background:#ffebee;color:#c62828}
.pill-high{background:#ffebee;color:#c62828}
.pill-med{background:#fff3e0;color:#e65100}
.pill-low{background:#e8f5e9;color:#2e7d32}
.cat-bars{margin-bottom:14px}
.cat-row{display:flex;align-items:center;gap:10px;margin-bottom:7px}
.cat-name{width:220px;font-size:11.5px;color:var(--text-mid);flex-shrink:0}
.bar-wrap{flex:1;background:#e8ecef;border-radius:4px;height:16px;overflow:hidden}
.bar-fill{height:100%;border-radius:4px;display:flex;align-items:center;padding-left:6px}
.bar-label{font-size:10px;color:white;font-weight:700;white-space:nowrap}
.cat-count{width:30px;text-align:right;font-size:12px;font-weight:700;color:var(--navy)}
details{margin-top:12px}
summary{cursor:pointer;font-size:11.5px;font-weight:600;color:var(--navy-light);padding:8px 12px;background:#f0f4f8;border-radius:6px;border:1px solid var(--border);user-select:none;list-style:none;display:flex;align-items:center;gap:8px}
summary::-webkit-details-marker{display:none}
summary::before{content:'▶';font-size:9px;transition:transform .2s;display:inline-block}
details[open] summary::before{transform:rotate(90deg)}
summary:hover{background:#e3eaf3}
.drill-body{padding:12px 0 0}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.divider{border:none;border-top:1px solid var(--border);margin:16px 0}
tr.total-row td{font-weight:700;background:var(--navy)!important;color:#fff;font-size:12.5px}
tr.alert-row td{background:#fff8e1!important;font-weight:600;color:#7f3700}
.sub-label{font-size:11px;font-weight:700;color:var(--navy);margin:14px 0 8px;text-transform:uppercase;letter-spacing:.5px;padding-bottom:4px;border-bottom:2px solid var(--navy);display:inline-block}
.age-pill{display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700}
.age-critical{background:#b71c1c;color:#fff}
.age-high{background:#e53935;color:#fff}
.age-med{background:#fb8c00;color:#fff}
.age-low{background:#66bb6a;color:#fff}
.footer{text-align:center;color:var(--text-light);font-size:11px;margin-top:24px;padding:12px;border-top:1px solid var(--border)}
"""


# ── MAIN HTML BUILDER ─────────────────────────────────────────────────────────

def build_html(tickets, fin, fin_label,
               vendors, vtotals,
               residential, commercial, collection,
               actions, commitments, commit_total):

    # Derived helpdesk stats
    ageing   = sorted([t for t in tickets if t["age"] >= 9], key=lambda x: -x["age"])
    inactive = sorted([t for t in tickets if t["idle"] >= 3 and t["status"] != "On Hold"],
                      key=lambda x: -x["idle"])
    on_hold  = [t for t in tickets if t["status"] == "On Hold"]
    cat_cnt  = Counter(t["category"] for t in tickets)

    # Action stats
    closed_n = sum(1 for a in actions if "Closed" in a["status"] or "✅" in a["status"])
    open_n   = sum(1 for a in actions if "Open"   in a["status"] or "🟡" in a["status"])

    # Financial
    total_cr = fin.get("total_cr", 0)
    total_dr = fin.get("total_dr", 0)
    deficit  = fin.get("deficit",  0)
    deficit_s  = f"&#8377;({ind(abs(deficit))})" if deficit < 0 else f"&#8377;{ind(deficit)}"
    deficit_cl = "red" if deficit < 0 else "green"

    # Outstanding vendors
    outstanding = [v for v in vendors if float(v.get("pending") or 0) > 0]
    outstanding_txt = " &nbsp;·&nbsp; ".join(
        f"{v['vendor']} &#8377;{ind(v['pending'])} ({v['purpose']})"
        for v in outstanding
    ) if outstanding else "All vendors fully paid"
    outstanding_cl = "red" if outstanding else "green"

    # Dues totals
    res_total = sum(r["total"] for r in residential)
    com_total = sum(c["total"] for c in commercial)

    # Labels
    now_label  = TODAY.strftime("%-d %b %Y")
    week_label = f"Week {WEEK_NUM} &nbsp;·&nbsp; {FY_LABEL}"

    res_rows_html, _ = res_due_rows(residential)
    com_rows_html, _ = com_due_rows(commercial)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TGT AOA — Management Dashboard</title>
<style>
{CSS}
</style>
</head>
<body>

<div class="hdr">
  <div>
    <div class="hdr-title">TGT AOA &#8212; Management Dashboard</div>
    <div class="hdr-sub">The Green Terraces &nbsp;·&nbsp; Electronic City Phase 1 &nbsp;·&nbsp; 393 Units</div>
  </div>
  <div class="hdr-meta">
    {week_label}<br>
    Data as on {now_label}<br>
    Prepared: Abhay Kapoor, Secretary
  </div>
</div>

<div class="page">

<!-- ═══════════════════ SECTION 1: HELPDESK ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#127917; Helpdesk &#8212; All Open Tickets
    <span class="sec-tag">Live Status &nbsp;·&nbsp; All Months</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-4">
      <div class="kpi amber"><div class="kpi-val">{len(tickets)}</div>
        <div class="kpi-lbl">Total Open (In Progress + Reopened + On Hold)</div></div>
      <div class="kpi red"><div class="kpi-val">{len(ageing)}</div>
        <div class="kpi-lbl">Ageing 9+ Days</div></div>
      <div class="kpi red"><div class="kpi-val">{len(inactive)}</div>
        <div class="kpi-lbl">Inactive 3+ Days (excl. On Hold)</div></div>
      <div class="kpi"><div class="kpi-val">{len(on_hold)}</div>
        <div class="kpi-lbl">On Hold</div></div>
    </div>
    <div class="alert amber">
      <strong>&#9888; Seepage Backlog &#8212; Monsoon Risk:</strong>
      {cat_cnt.get('Seepage', 0)} of {len(tickets)} open tickets are seepage issues,
      many inherited from previous committee. Monsoon season increases volume.
      Recommend a structured seepage closure drive &#8212; triage by root cause before peak.
    </div>
    {cat_bars_html(cat_cnt, len(tickets))}
    <details>
      <summary>&#128203; Ageing Tickets &#8212; Open 9+ Days ({len(ageing)} of {len(tickets)})</summary>
      <div class="drill-body"><div class="tbl-wrap tbl-scroll"><table>
        <thead><tr><th>Ticket #</th><th>Flat</th><th>Category</th><th>Subject</th>
          <th>Status</th><th class="r">Age</th><th>Created</th></tr></thead>
        <tbody>{ticket_table_rows(ageing, 'age')}</tbody>
      </table></div></div>
    </details>
    <details>
      <summary>&#128164; Inactive &#8212; No Update 3+ Days, excl. On Hold ({len(inactive)} of {len(tickets)})</summary>
      <div class="drill-body"><div class="tbl-wrap tbl-scroll"><table>
        <thead><tr><th>Ticket #</th><th>Flat</th><th>Category</th><th>Subject</th>
          <th>Status</th><th class="r">Idle</th><th>Last Update</th></tr></thead>
        <tbody>{ticket_table_rows(inactive, 'idle')}</tbody>
      </table></div></div>
    </details>
    <details>
      <summary>&#9208; On Hold Tickets ({len(on_hold)})</summary>
      <div class="drill-body"><div class="tbl-wrap"><table>
        <thead><tr><th>Ticket #</th><th>Flat</th><th>Category</th><th>Subject</th>
          <th class="r">Age</th></tr></thead>
        <tbody>{on_hold_rows(on_hold)}</tbody>
      </table></div></div>
    </details>
  </div>
</div>

<!-- ═══════════════════ SECTION 2: FINANCIALS ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#128176; Financial Summary &#8212; {fin_label} 2026
    <span class="sec-tag">Bank Statement &nbsp;·&nbsp; Current Month Only</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-3">
      <div class="kpi green"><div class="kpi-val">&#8377;{ind(total_cr)}</div>
        <div class="kpi-lbl">Total Income (CR)</div></div>
      <div class="kpi red"><div class="kpi-val">&#8377;{ind(total_dr)}</div>
        <div class="kpi-lbl">Total Expenses (DR)</div></div>
      <div class="kpi {deficit_cl}"><div class="kpi-val">{deficit_s}</div>
        <div class="kpi-lbl">Month Surplus / (Deficit)</div></div>
    </div>
    <div class="two-col">
      <div>
        <div class="sub-label">Income (CR)</div>
        <div class="tbl-wrap"><table>
          <thead><tr><th>Category</th><th class="r">&#8377;</th></tr></thead>
          <tbody>
            {_ordered_rows(fin.get('cr', {}), CR_ORDER)}
            <tr class="total-row"><td>Total CR</td><td class="r">{ind(total_cr)}</td></tr>
          </tbody>
        </table></div>
      </div>
      <div>
        <div class="sub-label">Expenses (DR)</div>
        <div class="tbl-wrap"><table>
          <thead><tr><th>Category</th><th class="r">&#8377;</th></tr></thead>
          <tbody>
            {_ordered_rows(fin.get('dr', {}), DR_ORDER)}
            <tr class="total-row"><td>Total DR</td><td class="r">{ind(total_dr)}</td></tr>
          </tbody>
        </table></div>
      </div>
    </div>
  </div>
</div>

<!-- ═══════════════════ SECTION 3: PAYMENTS ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#129534; Payment Status &#8212; {fin_label} 2026
    <span class="sec-tag">Vendor Bills &amp; Payments</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-4">
      <div class="kpi"><div class="kpi-val">&#8377;{ind(vtotals.get('payable', 0))}</div>
        <div class="kpi-lbl">Total Payable</div></div>
      <div class="kpi green"><div class="kpi-val">&#8377;{ind(vtotals.get('paid', 0))}</div>
        <div class="kpi-lbl">Total Paid</div></div>
      <div class="kpi {'amber' if vtotals.get('pending', 0) > 0 else 'green'}">
        <div class="kpi-val">&#8377;{ind(abs(vtotals.get('pending', 0)))}</div>
        <div class="kpi-lbl">Total Pending</div></div>
      <div class="kpi {outstanding_cl}"><div class="kpi-val">{len(outstanding)}</div>
        <div class="kpi-lbl">Vendors with Outstanding</div></div>
    </div>
    <div class="alert {'red' if outstanding else 'green'}">
      <strong>{'Outstanding:' if outstanding else 'Status:'}</strong> {outstanding_txt}
    </div>
    <details open>
      <summary>&#128209; Full Vendor Payment Detail</summary>
      <div class="drill-body"><div class="tbl-wrap"><table>
        <thead><tr><th>Vendor</th><th>Purpose</th>
          <th class="r">Billed (&#8377;)</th><th class="r">Payable after TDS (&#8377;)</th>
          <th class="r">Paid (&#8377;)</th><th class="r">Pending (&#8377;)</th></tr></thead>
        <tbody>
          {vendor_rows_html(vendors)}
          <tr class="total-row"><td colspan="2">TOTAL</td>
            <td class="r">{ind(sum(v['billed'] for v in vendors))}</td>
            <td class="r">{ind(vtotals.get('payable', 0))}</td>
            <td class="r">{ind(vtotals.get('paid', 0))}</td>
            <td class="r">{ind(abs(vtotals.get('pending', 0)))}</td>
          </tr>
        </tbody>
      </table></div></div>
    </details>
  </div>
</div>

<!-- ═══════════════════ SECTION 4: DUES & COLLECTION ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#128202; Pending Dues &amp; Collection Status
    <span class="sec-tag">Dues Report Raw &nbsp;·&nbsp; Collection Report Raw</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-4">
      <div class="kpi amber"><div class="kpi-val">{len(residential)}</div>
        <div class="kpi-lbl">Residential Units with Dues</div></div>
      <div class="kpi red"><div class="kpi-val">&#8377;{ind(res_total)}</div>
        <div class="kpi-lbl">Total Residential Dues</div></div>
      <div class="kpi"><div class="kpi-val">{len(commercial)}</div>
        <div class="kpi-lbl">Commercial Entities with Dues</div></div>
      <div class="kpi amber"><div class="kpi-val">&#8377;{ind(com_total)}</div>
        <div class="kpi-lbl">Total Commercial Dues</div></div>
    </div>
    <div class="sub-label">Residential Dues ({len(residential)} Units)</div>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Unit</th><th>Owner</th><th>Tenant</th>
        <th class="r">Penalty (&#8377;)</th><th class="r">Ledger Dues (&#8377;)</th>
        <th class="r">Total Dues (&#8377;)</th></tr></thead>
      <tbody>{res_rows_html}</tbody>
    </table></div>
    <hr class="divider">
    <div class="sub-label">Commercial / Non-Member Dues ({len(commercial)} Entities)</div>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Entity</th><th class="r">Ledger Dues (&#8377;)</th>
        <th class="r">Total Dues (&#8377;)</th></tr></thead>
      <tbody>{com_rows_html}</tbody>
    </table></div>
    <hr class="divider">
    <div class="sub-label">Collection Status</div>
    <div class="kpi-row kpi-row-4">
      {collection_kpis(collection)}
    </div>
  </div>
</div>

<!-- ═══════════════════ SECTION 5: ACTION TRACKER ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#9989; Action Tracker &#8212; {FY_LABEL}
    <span class="sec-tag">All Actions</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-4">
      <div class="kpi"><div class="kpi-val">{len(actions)}</div>
        <div class="kpi-lbl">Total Actions Logged</div></div>
      <div class="kpi green"><div class="kpi-val">{closed_n}</div>
        <div class="kpi-lbl">Closed</div></div>
      <div class="kpi {'amber' if open_n > 0 else 'green'}">
        <div class="kpi-val">{open_n}</div>
        <div class="kpi-lbl">Open</div></div>
      <div class="kpi"><div class="kpi-val">0</div>
        <div class="kpi-lbl">Overdue</div></div>
    </div>
    <div class="tbl-wrap"><table>
      <thead><tr><th>#</th><th>Week</th><th>Category</th><th>Description</th>
        <th>Assigned To</th><th>Priority</th><th class="c">Status</th></tr></thead>
      <tbody>{action_rows_html(actions)}</tbody>
    </table></div>
  </div>
</div>

<!-- ═══════════════════ SECTION 6: COMMITMENTS ═══════════════════ -->
<div class="section">
  <div class="sec-hdr">&#128203; Approved Commitments &#8212; Pending Bill
    <span class="sec-tag">Approved, Bill Not Yet Received</span></div>
  <div class="sec-body">
    <div class="kpi-row kpi-row-3">
      <div class="kpi"><div class="kpi-val">{len(commitments)}</div>
        <div class="kpi-lbl">Items Approved, Bill Awaited</div></div>
      <div class="kpi amber"><div class="kpi-val">&#8377;{ind(commit_total)}</div>
        <div class="kpi-lbl">Estimated Liability</div></div>
      <div class="kpi"><div class="kpi-val">{fin_label} 2026</div>
        <div class="kpi-lbl">Expected Billing Month</div></div>
    </div>
    <div class="tbl-wrap"><table>
      <thead><tr><th>#</th><th>Date Approved</th><th>Vendor</th><th>Description</th>
        <th>Category</th><th class="r">Est. Amount (&#8377;)</th><th>Billing Month</th></tr></thead>
      <tbody>
        {commitment_rows_html(commitments)}
        <tr class="total-row"><td colspan="5">TOTAL ESTIMATED LIABILITY</td>
          <td class="r">{ind(commit_total)}</td><td></td></tr>
      </tbody>
    </table></div>
  </div>
</div>

<div class="footer">
  TGT AOA Management Dashboard &nbsp;·&nbsp; {week_label} &nbsp;·&nbsp;
  Data as on {now_label} &nbsp;·&nbsp; Prepared by Abhay Kapoor, Secretary
</div>

</div><!-- /page -->
</body>
</html>"""


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if not MASTER_FILE.exists():
        print(f"ERROR: Master File not found:\n  {MASTER_FILE}\n")
        sys.exit(1)

    print("Reading Master File...")
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)

    tickets             = get_tickets(wb)
    fin, fin_label      = get_financials(wb)
    vendors, vtotals    = get_payments(wb)
    residential, commercial = get_dues(wb)
    coll                = get_collection(wb)
    actions             = get_actions(wb)
    commitments, ctotal = get_commitments(wb)

    print(f"  Open tickets  : {len(tickets)}")
    print(f"  Financial month: {fin_label}")
    print(f"  Vendors       : {len(vendors)}")
    print(f"  Res dues      : {len(residential)} units")
    print(f"  Com dues      : {len(commercial)} entities")
    print(f"  Collection    : {len(coll)} months")
    print(f"  Actions       : {len(actions)}")
    print(f"  Commitments   : {len(commitments)} pending")

    html = build_html(
        tickets, fin, fin_label,
        vendors, vtotals,
        residential, commercial, coll,
        actions, commitments, ctotal,
    )

    OUTPUT_FILE.write_text(html, encoding="utf-8")
    print(f"\n✅  Generated: {OUTPUT_FILE}")
    print("    Next step : git add index.html && git commit -m 'Dashboard update' && git push")
